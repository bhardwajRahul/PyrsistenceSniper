"""Tests for the console, CSV, HTML and XLSX report writers."""

from __future__ import annotations

import csv
import io
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook
from pyrsistencesniper.core.models import (
    AccessLevel,
    CheckFailure,
    Enrichment,
    Finding,
    HiveRecord,
    HiveStatus,
)
from pyrsistencesniper.output.base import label_for, sanitize_cell
from pyrsistencesniper.output.console_output import ConsoleOutput
from pyrsistencesniper.output.csv_output import CsvOutput
from pyrsistencesniper.output.html_output import HtmlOutput
from pyrsistencesniper.output.xlsx_output import XlsxOutput


def _make_result(
    path: str = "HKLM\\Run", value: str = "test.exe"
) -> tuple[Finding, tuple[Enrichment, ...]]:
    """Build one result pair with every core Finding field set and no enrichments."""
    finding = Finding(
        path=path,
        value=value,
        technique="Test",
        mitre_id="T0000",
        description="Test description",
        access_gained=AccessLevel.SYSTEM,
        hostname="HOST",
        check_id="test_check",
    )
    return (finding, ())


def _dated_finding() -> Finding:
    """Build a finding the timeline stage already dated from the $MFT."""
    return Finding(
        path="HKLM\\Run",
        value="test.exe",
        check_id="test_check",
        last_change="2026-07-20 12:00:00",
        change_source="$MFT",
    )


def _hive(name: str, status: HiveStatus, **kwargs: object) -> HiveRecord:
    """Build a hive record for the scan-integrity assertions."""
    return HiveRecord(name=name, status=status, **kwargs)  # type: ignore[arg-type]


def _html(results: list[Any], **kwargs: Any) -> str:
    """Render the HTML report and return the artifact text."""
    out = io.StringIO()
    HtmlOutput()._write(results, out, **kwargs)
    return out.getvalue()


def _csv(results: list[Any], **kwargs: Any) -> str:
    """Render the CSV report and return the artifact text."""
    out = io.StringIO()
    CsvOutput()._write(results, out, **kwargs)
    return out.getvalue()


def _console(results: list[Any], **kwargs: Any) -> str:
    """Render the console report and return the text an analyst reads."""
    out = io.StringIO()
    ConsoleOutput()._write(results, out, **kwargs)
    return out.getvalue()


def _workbook(results: list[Any], output: Any = None, **kwargs: Any) -> Workbook:
    """Render the workbook and load it back; the default target is held in memory."""
    target = io.BytesIO() if output is None else output
    XlsxOutput().render(results, output=target, **kwargs)
    if isinstance(target, io.BytesIO):
        target.seek(0)
    return load_workbook(target)


def _headers(sheet: Any) -> list[str]:
    """Return the header labels of a rendered sheet."""
    return [cell.value for cell in sheet[1]]


def _cell_under(sheet: Any, label: str) -> Any:
    """Return the first data cell of the column carrying the given header label."""
    return sheet.cell(row=2, column=_headers(sheet).index(label) + 1)


@pytest.mark.parametrize(
    ("raw", "expected"),
    [
        ("=CMD()", "'=CMD()"),
        ("+1+2", "'+1+2"),
        ("-1-2", "'-1-2"),
        ("@SUM(A1)", "'@SUM(A1)"),
        ("\t=CMD()", "'\t=CMD()"),
        ("\r=CMD()", "'\r=CMD()"),
        ("\n=CMD()", "'\n=CMD()"),
        (" =CMD()", "' =CMD()"),
        ("explorer.exe", "explorer.exe"),
        ("", ""),
        (42, "42"),
    ],
)
def test_sanitize_cell_formula_prefixes(raw: object, expected: str) -> None:
    """Every formula-trigger prefix is quoted; ordinary values pass through."""
    assert sanitize_cell(raw) == expected


def test_sanitize_cell_replaces_illegal_control_characters() -> None:
    """Control characters rejected by the spreadsheet formats are replaced."""
    assert "\x01" not in sanitize_cell("C:\\evil\x01.exe")


def test_label_for_core_field() -> None:
    """Core fields keep their Finding.FIELDS labels."""
    assert label_for("mitre_id") == "MITRE ID"


def test_label_for_enrichment_field() -> None:
    """Enrichment keys render as title-cased provider and key."""
    assert label_for("enrichment.triage.notes") == "Triage Notes"


def test_label_for_unknown_field_passthrough() -> None:
    """Unknown fields fall back to the raw name."""
    assert label_for("custom_column") == "custom_column"


def test_html_labels_enrichment_column() -> None:
    """The HTML report labels enrichment columns, not their raw dotted keys."""
    finding = Finding(path="HKLM\\Run", value="test.exe", check_id="test_check")
    enrichment = Enrichment(provider="triage", data={"notes": "check it"})
    assert "Triage Notes" in _html([(finding, (enrichment,))])


def test_html_autoescaping_value() -> None:
    """A <script> tag in value must be encoded as \\u003c in JSON payload."""
    html = _html([_make_result(value="<script>alert(1)</script>")])
    assert "<script>alert(1)</script>" not in html
    assert "\\u003cscript" in html


def test_html_autoescaping_path() -> None:
    """An XSS payload in path must be encoded as \\u003c in JSON payload."""
    html = _html([_make_result(path='HKLM\\<img src=x onerror="alert(1)">')])
    assert "<img src=x" not in html
    assert "\\u003cimg" in html


def test_csv_output_sanitizes_all_fields() -> None:
    """A formula trigger from the registry is quoted before a spreadsheet runs it."""
    assert "'=HYPERLINK()" in _csv([_make_result(value="=HYPERLINK()")])


def test_xlsx_produces_valid_workbook() -> None:
    """The workbook opens with labelled columns, so a value is read under its header."""
    sheet = _workbook([_make_result(value="malware.exe")]).active
    assert "Path" in _headers(sheet)
    assert _cell_under(sheet, "Value").value == "malware.exe"


def test_xlsx_raises_without_output() -> None:
    """A workbook has no console form, so a missing target fails loudly not silently."""
    with pytest.raises(ValueError, match="requires a file path"):
        XlsxOutput().render([_make_result()])


def test_xlsx_neutralizes_formula_injection() -> None:
    """A value starting with = is quoted so the workbook holds no live formula."""
    sheet = _workbook([_make_result(value="=cmd|'/c calc.exe'!A0")]).active
    cell = _cell_under(sheet, "Value")
    assert cell.data_type != "f"
    assert str(cell.value).startswith("'=")


def test_xlsx_survives_illegal_control_character() -> None:
    """A control character in a value still yields a readable workbook."""
    sheet = _workbook([_make_result(value="C:\\evil\x01payload.exe")]).active
    assert "\x01" not in str(_cell_under(sheet, "Value").value)


def test_xlsx_enrichment_columns() -> None:
    """Provider data reaches the workbook as its own column, not as a lost field."""
    enrichment = Enrichment(provider="vt", data={"score": "5/70"})
    finding, _ = _make_result()
    sheet = _workbook([(finding, (enrichment,))]).active
    assert "Vt Score" in _headers(sheet)
    assert _cell_under(sheet, "Vt Score").value == "5/70"


def test_xlsx_empty_results() -> None:
    """A scan that found nothing still yields a readable workbook, not an empty file."""
    sheet = _workbook([]).active
    assert sheet.cell(row=1, column=1).value == "Path"
    assert sheet.cell(row=2, column=1).value is None


def test_html_renders_change_candidates() -> None:
    """Change candidates ride in the DATA payload as _change_candidates."""
    finding = Finding(
        path="HKLM\\Run",
        value="test.exe",
        check_id="test_check",
        change_candidates=("2026-07-20 12:00:00 - event log - x",),
    )
    assert "2026-07-20 12:00:00 - event log - x" in _html([(finding, ())])


def test_html_includes_column_help() -> None:
    """The report embeds COLUMN_HELP carrying the Last Change guidance text."""
    finding = Finding(path="HKLM\\Run", value="test.exe", check_id="test_check")
    html = _html([(finding, ())])
    assert "COLUMN_HELP" in html
    assert "survives collection" in html


def test_html_renders_change_columns() -> None:
    """last_change and change_source populate the row payload."""
    html = _html([(_dated_finding(), ())])
    assert "2026-07-20 12:00:00" in html
    assert "$MFT" in html


def test_csv_includes_change_columns() -> None:
    """CSV header and rows carry the two timeline columns."""
    csv_text = _csv([(_dated_finding(), ())])
    assert "Last Change" in csv_text
    assert "Change Source" in csv_text
    assert "2026-07-20 12:00:00" in csv_text


def test_xlsx_includes_change_columns() -> None:
    """XLSX header carries the Last Change and Change Source columns with values."""
    sheet = _workbook([(_dated_finding(), ())]).active
    assert "Change Source" in _headers(sheet)
    assert _cell_under(sheet, "Last Change").value == "2026-07-20 12:00:00"


def test_console_reports_an_unreadable_hive() -> None:
    """An empty console report says which hive was never read."""
    text = _console(
        [],
        inventory=(_hive("SYSTEM", HiveStatus.OPEN_FAILED, error="OSError: bad hive"),),
    )

    assert "SCAN INCOMPLETE" in text
    assert "SYSTEM" in text
    assert "OSError: bad hive" in text
    assert "No findings." in text


def test_console_stays_quiet_when_every_hive_was_read() -> None:
    """A clean scan gets no integrity banner."""
    assert "SCAN INCOMPLETE" not in _console(
        [], inventory=(_hive("SOFTWARE", HiveStatus.OPENED),)
    )


def test_console_reports_a_failed_check() -> None:
    """An empty console report names the check that failed, not just an empty result."""
    text = _console(
        [], failures=(CheckFailure(check_id="run_keys", error="OSError: bad cell"),)
    )

    assert "SCAN INCOMPLETE" in text
    assert "run_keys" in text
    assert "OSError: bad cell" in text


def test_console_stays_quiet_when_every_check_ran() -> None:
    """A scan where nothing failed gets no failed-check banner."""
    assert "did not run" not in _console([], failures=())


def test_html_renders_the_failed_check_block() -> None:
    """The HTML artifact carries failed checks, so it is not read as clean."""
    html = _html(
        [(Finding(path="p", value="v"), ())],
        failures=(CheckFailure(check_id="com_hijack", error="OSError: bad cell"),),
    )

    assert "did not run" in html
    assert "com_hijack" in html


def test_xlsx_writes_a_failed_check_sheet(tmp_path: Path) -> None:
    """The workbook carries a sheet naming every check that produced nothing."""
    workbook = _workbook(
        [(Finding(path="p", value="v"), ())],
        tmp_path / "report.xlsx",
        failures=(CheckFailure(check_id="ghost_task", error="ValueError: boom"),),
    )

    assert "Failed Checks" in workbook.sheetnames
    rows = list(workbook["Failed Checks"].iter_rows(values_only=True))
    assert rows[1] == ("ghost_task", "ValueError: boom")


def test_html_renders_the_unreadable_hive_block() -> None:
    """The HTML report carries the integrity warning into the artifact."""
    html = _html(
        [(Finding(path="p", value="v"), ())],
        inventory=(_hive("SYSTEM", HiveStatus.OPEN_FAILED, error="OSError: bad hive"),),
    )

    assert "Scan incomplete" in html
    assert "SYSTEM" in html


def test_html_omits_the_block_when_every_hive_was_read() -> None:
    """A clean report carries no permanent scare banner."""
    assert "Scan incomplete" not in _html(
        [(Finding(path="p", value="v"), ())],
        inventory=(_hive("SOFTWARE", HiveStatus.OPENED),),
    )


def test_html_notes_dirty_hives_separately() -> None:
    """A hive that was read but is dirty is a caveat, not a failure."""
    html = _html(
        [(Finding(path="p", value="v"), ())],
        inventory=(_hive("SOFTWARE", HiveStatus.OPENED, dirty=True),),
    )

    assert "uncommitted transactions" in html
    assert "Scan incomplete" not in html


def test_html_filters_every_column_rather_than_a_fixed_four() -> None:
    """Filtering is per column, so a new field is filterable without a code change."""
    # Asserted at string level because the machinery only runs in a browser; the
    # names checked absent are the hardcoded severity, technique, MITRE ID and
    # access-level controls this replaced.
    html = _html([(Finding(path="HKLM\\Run", value="a.exe"), ())])

    assert "col-filter-btn" in html
    assert "openFilterMenu" in html
    assert "columnFilters" in html
    for removed in ("btn-filter-severity", "btn-filter-technique"):
        assert removed not in html
    for removed_label in ("All Severities", "All MITRE IDs"):
        assert removed_label not in html


def test_html_shows_active_filters_as_chips() -> None:
    """An active filter is visible without opening the menu that set it."""
    html = _html([(Finding(path="HKLM\\Run", value="a.exe"), ())])

    assert 'id="filter-chips"' in html
    assert "renderChips" in html


def test_csv_ignores_the_inventory() -> None:
    """The CSV stays a pure findings table so downstream parsers keep working."""
    csv_text = _csv(
        [(Finding(path="p", value="v"), ())],
        inventory=(_hive("SYSTEM", HiveStatus.OPEN_FAILED),),
    )
    rows = list(csv.reader(io.StringIO(csv_text)))

    assert len(rows) == 2
    assert "SYSTEM" not in csv_text


def test_xlsx_appends_a_hive_sheet(tmp_path: Path) -> None:
    """The workbook gains a Hives sheet without disturbing the findings grid."""
    workbook = _workbook(
        [(Finding(path="p", value="v"), ())],
        tmp_path / "report.xlsx",
        inventory=(_hive("SYSTEM", HiveStatus.OPEN_FAILED, error="OSError: bad"),),
    )

    assert workbook.sheetnames == ["Findings", "Hives"]
    assert workbook.active.title == "Findings"
    assert workbook.active.cell(row=1, column=1).value == "Path"
    assert workbook["Hives"].cell(row=2, column=1).value == "SYSTEM"


def test_xlsx_without_inventory_has_only_the_findings_sheet(tmp_path: Path) -> None:
    """The default keeps the workbook byte-compatible with existing consumers."""
    workbook = _workbook([(Finding(path="p", value="v"), ())], tmp_path / "report.xlsx")

    assert workbook.sheetnames == ["Findings"]
