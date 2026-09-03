"""XLSX workbook renderer for spreadsheet based triage."""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import IO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from pyrsistencesniper.core.models import AnnotatedResult, CheckFailure, HiveRecord
from pyrsistencesniper.output.base import OutputBase, label_for, sanitize_cell

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_MAX_COLUMN_WIDTH = 60


def _write_header_row(sheet: Worksheet, labels: Iterable[str]) -> None:
    """Write a styled, frozen header row across the top of a sheet."""
    for column_index, label in enumerate(labels, start=1):
        cell = sheet.cell(row=1, column=column_index, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    sheet.freeze_panes = "A2"


def _write_rows(sheet: Worksheet, rows: Iterable[Sequence[object]]) -> None:
    """Write sanitized data rows beneath the header row."""
    for row_index, values in enumerate(rows, start=2):
        for column_index, value in enumerate(values, start=1):
            sheet.cell(row=row_index, column=column_index, value=sanitize_cell(value))


class XlsxOutput(OutputBase):
    """Writes findings as a styled XLSX workbook."""

    def render(
        self,
        results: list[AnnotatedResult],
        output: Path | IO[bytes] | None = None,  # type: ignore[override]
        *,
        inventory: tuple[HiveRecord, ...] = (),
        failures: tuple[CheckFailure, ...] = (),
    ) -> None:
        if output is None:
            msg = "XLSX output requires a file path or binary stream"
            raise ValueError(msg)

        rows, fieldnames = self._flatten_results(results)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Findings"
        _write_header_row(sheet, (label_for(name) for name in fieldnames))
        _write_rows(sheet, ([row.get(name, "") for name in fieldnames] for row in rows))
        self._fit_columns(sheet, fieldnames, len(rows))

        if inventory:
            self._write_hive_sheet(workbook, inventory)
        if failures:
            self._write_failure_sheet(workbook, failures)

        if isinstance(output, Path):
            workbook.save(str(output))
        else:
            workbook.save(output)

    @staticmethod
    def _fit_columns(sheet: Worksheet, fieldnames: list[str], row_count: int) -> None:
        """Widen each column to its longest cell, up to a readable maximum."""
        for column_index, name in enumerate(fieldnames, start=1):
            widest = len(label_for(name))
            for row_index in range(2, row_count + 2):
                value = sheet.cell(row=row_index, column=column_index).value
                if value is not None:
                    widest = max(widest, len(str(value)))
            letter = get_column_letter(column_index)
            sheet.column_dimensions[letter].width = min(widest + 2, _MAX_COLUMN_WIDTH)

    @staticmethod
    def _write_hive_sheet(
        workbook: Workbook, inventory: tuple[HiveRecord, ...]
    ) -> None:
        """Append a sheet naming every hive the scan expected and what it got."""
        sheet = workbook.create_sheet("Hives")
        _write_header_row(sheet, ("Hive", "User", "Status", "Dirty", "Path", "Error"))
        _write_rows(
            sheet,
            [
                (
                    record.name,
                    record.owner,
                    record.status.value,
                    "yes" if record.dirty else "",
                    record.path,
                    record.error,
                )
                for record in inventory
            ],
        )

    @staticmethod
    def _write_failure_sheet(
        workbook: Workbook, failures: tuple[CheckFailure, ...]
    ) -> None:
        """Append a sheet naming every check that raised and produced no findings."""
        sheet = workbook.create_sheet("Failed Checks")
        _write_header_row(sheet, ("Check", "Error"))
        _write_rows(sheet, [(failure.check_id, failure.error) for failure in failures])

    def _write(
        self,
        results: list[AnnotatedResult],
        out: IO[str],
        *,
        inventory: tuple[HiveRecord, ...] = (),
        failures: tuple[CheckFailure, ...] = (),
    ) -> None:
        msg = "XLSX is binary; use render() directly"
        raise NotImplementedError(msg)
